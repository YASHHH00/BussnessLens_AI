"""
BusinessLens AI — Report Generator Engine

Generates downloadable spreadsheet reports (Excel workbook via openpyxl)
or structured export bundles containing dashboard KPIs and metric data.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ReportGeneratorEngine:
    def build_excel_report(
        self,
        title: str,
        kpis: list[dict[str, Any]],
        data_rows: list[dict[str, Any]],
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive BI Summary"

        # Styling
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Title
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value = title
        cell.font = title_font
        cell.fill = title_fill

        # KPI Section
        ws["A3"] = "Key Performance Indicators"
        ws["A3"].font = Font(size=12, bold=True)

        ws["A4"] = "KPI Name"
        ws["B4"] = "Value"
        ws["C4"] = "Unit"
        ws["A4"].font = header_font
        ws["B4"].font = header_font
        ws["C4"].font = header_font
        ws["A4"].fill = header_fill
        ws["B4"].fill = header_fill
        ws["C4"].fill = header_fill

        row_idx = 5
        for kpi in kpis:
            ws[f"A{row_idx}"] = kpi.get("display_name", kpi.get("name", ""))
            ws[f"B{row_idx}"] = kpi.get("formatted_value", kpi.get("val", ""))
            ws[f"C{row_idx}"] = kpi.get("unit", "")
            row_idx += 1

        # Data rows section
        row_idx += 2
        ws[f"A{row_idx}"] = "Detailed Data"
        ws[f"A{row_idx}"].font = Font(size=12, bold=True)
        row_idx += 1

        if data_rows:
            cols = list(data_rows[0].keys())
            for c_idx, col_name in enumerate(cols, start=1):
                c = ws.cell(row=row_idx, column=c_idx, value=col_name)
                c.font = header_font
                c.fill = header_fill
            row_idx += 1

            for row_data in data_rows:
                for c_idx, col_name in enumerate(cols, start=1):
                    ws.cell(row=row_idx, column=c_idx, value=row_data.get(col_name))
                row_idx += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
